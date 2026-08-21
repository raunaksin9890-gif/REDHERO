import csv
import io
import os
from datetime import date, datetime, timedelta
from uuid import uuid4

from django.conf import settings
from django.http import HttpResponse
from mongoengine.errors import NotUniqueError, ValidationError
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response

from .models import (
    Assignment,
    AssignmentSubmission,
    Attendance,
    Blog,
    CLASSES,
    CurrentAffair,
    DEFAULT_PASSWORD,
    Exam,
    ExamAnswer,
    ExamAttempt,
    ExamQuestion,
    Fee,
    FeePayment,
    Marks,
    Note,
    NoteBookmark,
    Notice,
    PracticeAnswer,
    PracticeSession,
    QuestionBankQuestion,
    ROLE_ADMIN,
    ROLE_STUDENT,
    ROLE_TEACHER,
    Student,
    StudentMistake,
    StudyPlan,
    StudyPlanTask,
    Teacher,
    Timetable,
    TimetablePeriod,
    TopicPerformance,
    User,
    Video,
)
from .notifications import notify_all_students, notify_student, notify_students_for_class
from .security import current_user, hash_password, require_roles
from .services import next_code, parse_date, schedule_now, schedule_time, store_schedule_time


def ok(data=None, http_status=status.HTTP_200_OK):
    return Response(data or {}, status=http_status)


def bad(message, http_status=status.HTTP_400_BAD_REQUEST):
    return Response({"detail": message}, status=http_status)


def oid(value):
    return str(value.id) if value else None


def dt(value):
    return value.isoformat() if value else None


def user_json(user):
    return {
        "id": oid(user),
        "name": user.name,
        "email": user.email,
        "role": user.role,
        "force_password_change": user.force_password_change,
    } if user else None


def student_json(student):
    return {
        "id": oid(student),
        "user_id": oid(student.user),
        "student_id": student.student_id,
        "name": student.name,
        "email": student.email,
        "class_level": student.class_level,
        "division": student.division,
        "roll_number": student.roll_number,
        "profile_photo": student.profile_photo,
    }


def teacher_json(teacher):
    return {
        "id": oid(teacher),
        "user_id": oid(teacher.user),
        "teacher_id": teacher.teacher_id,
        "name": teacher.name,
        "email": teacher.email,
        "subjects": teacher.subjects,
        "assigned_classes": teacher.assigned_classes,
    }


def simple_json(row, fields):
    data = {"id": oid(row)}
    for field in fields:
        value = getattr(row, field)
        data[field] = dt(value) if isinstance(value, datetime) else value
    return data


def get_student_for_user(user):
    return Student.objects(user=user).first()


def get_teacher_for_user(user):
    return Teacher.objects(user=user).first()


def teacher_classes(user):
    teacher = get_teacher_for_user(user)
    return teacher.assigned_classes if teacher else []


def teacher_subjects(user):
    teacher = get_teacher_for_user(user)
    return teacher.subjects if teacher else []


def enforce_teacher_class(user, class_level):
    if user.role == ROLE_TEACHER and str(class_level) not in teacher_classes(user):
        raise PermissionDenied("Teachers can only access assigned classes")


def enforce_teacher_subject(user, subject):
    subjects = [item.lower() for item in teacher_subjects(user)]
    if user.role == ROLE_TEACHER and subjects and str(subject).lower() not in subjects:
        raise PermissionDenied("Teachers can only access assigned subjects")


def enforce_teacher_question_scope(user, class_level, subject):
    enforce_teacher_class(user, class_level)
    enforce_teacher_subject(user, subject)


def enforce_teacher_student(user, student):
    if user.role == ROLE_TEACHER and (not student or student.class_level not in teacher_classes(user)):
        raise PermissionDenied("Teachers can only access assigned classes")


def enforce_owner(user, row, field):
    if user.role == ROLE_TEACHER and getattr(row, field) != user:
        raise PermissionDenied("Teachers can only manage their own content")


def attendance_lock_collection():
    return Attendance._get_db()["AttendanceLocks"]


def attendance_audit_collection():
    return Attendance._get_db()["AttendanceAuditLogs"]


def validate_password_strength(password):
    if len(password) < 8:
        return "Password must be at least 8 characters"
    if not any(char.isupper() for char in password):
        return "Password must include an uppercase letter"
    if not any(char.islower() for char in password):
        return "Password must include a lowercase letter"
    if not any(char.isdigit() for char in password):
        return "Password must include a number"
    if not any(not char.isalnum() for char in password):
        return "Password must include a special character"
    return ""


def is_locked(row):
    if not row:
        return False
    return bool(attendance_lock_collection().find_one({"attendance_id": oid(row), "locked": True}))


def set_locked(row, user, locked):
    now = datetime.utcnow()
    attendance_lock_collection().update_one(
        {"attendance_id": oid(row)},
        {
            "$set": {
                "attendance_id": oid(row),
                "student_id": oid(row.student),
                "student_name": row.student.name if row.student else "",
                "class_level": row.class_level,
                "date": row.date,
                "locked": locked,
                "updated_by": oid(user),
                "updated_by_name": user.name,
                "updated_at": now,
            },
            "$setOnInsert": {"created_at": now},
        },
        upsert=True,
    )
    log_attendance("lock" if locked else "unlock", user, row)


def log_attendance(action, user, row, before=None, after=None):
    attendance_audit_collection().insert_one(
        {
            "attendance_id": oid(row),
            "student_id": oid(row.student),
            "student_name": row.student.name if row.student else "",
            "class_level": row.class_level,
            "date": row.date,
            "action": action,
            "before": before or {},
            "after": after or {},
            "performed_by": oid(user),
            "performed_by_name": user.name,
            "performed_by_role": user.role,
            "created_at": datetime.utcnow(),
        }
    )


def attendance_json(row):
    return {
        "id": oid(row),
        "student": student_json(row.student) if row.student else None,
        "class_level": row.class_level,
        "date": dt(row.date),
        "status": row.status,
        "marked_by": user_json(row.marked_by),
        "created_at": dt(row.created_at),
        "updated_at": dt(getattr(row, "updated_at", None)),
        "locked": is_locked(row),
    }


def marks_json(row):
    percent = round((row.marks_obtained / row.max_marks) * 100, 2) if row.max_marks else 0
    return {
        "id": oid(row),
        "student": student_json(row.student) if row.student else None,
        "class_level": row.class_level,
        "subject": row.subject,
        "exam_type": row.exam_type,
        "marks_obtained": row.marks_obtained,
        "max_marks": row.max_marks,
        "percentage": percent,
    }


def notice_json(row):
    return {
        "id": oid(row),
        "title": row.title,
        "body": row.body,
        "class_level": row.class_level,
        "created_at": dt(row.created_at),
    }


def timetable_json(row):
    return {"id": oid(row), "class_level": row.class_level, "periods": [period.to_mongo().to_dict() for period in row.periods]}


def assignment_submission_json(submission, include_student=False):
    data = {
        "id": oid(submission),
        "answer_text": submission.answer_text,
        "file_url": submission.file_url,
        "status": submission.status,
        "submitted_at": dt(submission.submitted_at),
        "updated_at": dt(getattr(submission, "updated_at", None)),
    }
    if include_student:
        data["student"] = student_json(submission.student) if submission.student else None
    return data


def assignment_status_for(row, student=None):
    submission = AssignmentSubmission.objects(assignment=row, student=student).first() if student else None
    if submission:
        return "Completed" if submission.status == "reviewed" else "Submitted"
    return "Overdue" if row.deadline and row.deadline < datetime.utcnow() else "Pending"


def assignment_manager_status(row, submission_count):
    if submission_count:
        return "Submitted"
    return "Overdue" if row.deadline and row.deadline < datetime.utcnow() else "Pending"


def assignment_json(row, user):
    submissions = list(AssignmentSubmission.objects(assignment=row).order_by("-submitted_at"))
    data = {
        "id": oid(row),
        "title": row.title,
        "description": row.description,
        "class_level": row.class_level,
        "subject": row.subject,
        "deadline": dt(row.deadline),
        "file_url": getattr(row, "file_url", ""),
        "created_at": dt(row.created_at),
        "updated_at": dt(getattr(row, "updated_at", None)),
        "created_by": user_json(row.created_by),
        "submission_count": len(submissions),
    }
    if user.role == ROLE_STUDENT:
        student = get_student_for_user(user)
        own = AssignmentSubmission.objects(assignment=row, student=student).first() if student else None
        data["own_submission"] = assignment_submission_json(own) if own else None
        data["status"] = assignment_status_for(row, student)
    else:
        data["submissions"] = [assignment_submission_json(item, include_student=True) for item in submissions]
        data["status"] = assignment_manager_status(row, len(submissions))
    return data


def assignment_payload(data, user, row=None):
    class_level = str(data.get("class_level", row.class_level if row else "")).strip()
    subject = str(data.get("subject", row.subject if row else "")).strip()
    title = str(data.get("title", row.title if row else "")).strip()
    description = str(data.get("description", row.description if row else "")).strip()
    if class_level not in CLASSES:
        raise ValueError("Select a valid class")
    if not title:
        raise ValueError("Assignment title is required")
    if not subject:
        raise ValueError("Subject is required")
    if not description:
        raise ValueError("Description is required")
    if not row and not data.get("deadline"):
        raise ValueError("Due date and due time are required")
    enforce_teacher_class(user, class_level)
    enforce_teacher_subject(user, subject)
    return {
        "title": title,
        "description": description,
        "class_level": class_level,
        "subject": subject,
        "deadline": parse_date(data.get("deadline", row.deadline if row else None)),
        "file_url": str(data.get("file_url", getattr(row, "file_url", "") if row else "") or "").strip(),
        "updated_at": datetime.utcnow(),
    }


def scoped_students(user):
    if user.role == ROLE_TEACHER:
        return Student.objects(class_level__in=teacher_classes(user))
    if user.role == ROLE_STUDENT:
        return Student.objects(user=user)
    return Student.objects


@api_view(["GET"])
def users(request):
    require_roles(request, [ROLE_ADMIN])
    return ok({"results": [user_json(user) for user in User.objects.order_by("-created_at")]})


@api_view(["POST"])
def reset_password(request, user_id):
    require_roles(request, [ROLE_ADMIN])
    target = User.objects(id=user_id).first()
    if not target:
        return bad("User not found", status.HTTP_404_NOT_FOUND)
    new_password = request.data.get("new_password") or DEFAULT_PASSWORD
    confirm_password = request.data.get("confirm_password", new_password)
    if new_password != confirm_password:
        return bad("Confirm password does not match")
    if new_password != DEFAULT_PASSWORD:
        strength_error = validate_password_strength(new_password)
        if strength_error:
            return bad(strength_error)
    force_change = bool(request.data.get("force_password_change", new_password == DEFAULT_PASSWORD))
    target.update(password_hash=hash_password(new_password), first_login=force_change, force_password_change=force_change, updated_at=datetime.utcnow())
    if new_password == DEFAULT_PASSWORD:
        return ok({"message": "Password reset to default", "default_password": DEFAULT_PASSWORD})
    return ok({"message": "Password updated successfully"})


@api_view(["POST"])
def force_password_change(request, user_id):
    require_roles(request, [ROLE_ADMIN])
    target = User.objects(id=user_id).first()
    if not target:
        return bad("User not found", status.HTTP_404_NOT_FOUND)
    target.update(force_password_change=True, updated_at=datetime.utcnow())
    return ok({"message": "Password change required on next login"})


@api_view(["GET"])
def dashboard(request):
    user = current_user(request)
    if user.role == ROLE_TEACHER:
        assigned = teacher_classes(user)
        return ok(
            {
                "assigned_classes": assigned,
                "students": Student.objects(class_level__in=assigned).count(),
                "recent_notices": [notice_json(row) for row in Notice.objects(class_level__in=assigned + ["all"]).order_by("-created_at")[:5]],
            }
        )
    if user.role == ROLE_STUDENT:
        student = get_student_for_user(user)
        attendance_rows = list(Attendance.objects(student=student))
        present = len([row for row in attendance_rows if row.status == "present"])
        return ok(
            {
                "profile": student_json(student),
                "attendance_percentage": round((present / len(attendance_rows)) * 100, 2) if attendance_rows else 0,
                "marks": [marks_json(row) for row in Marks.objects(student=student).order_by("-created_at")[:5]],
                "latest_notices": [notice_json(row) for row in Notice.objects(class_level__in=[student.class_level, "all"]).order_by("-created_at")[:5]],
                "current_affairs": [simple_json(row, ["title", "summary", "category", "image_url", "published_on"]) for row in CurrentAffair.objects.order_by("-published_on")[:4]],
                "recent_videos": [simple_json(row, ["title", "class_level", "subject", "chapter", "youtube_url"]) for row in Video.objects(class_level=student.class_level).order_by("-created_at")[:4]],
            }
        )
    return ok(
        {
            "total_students": Student.objects.count(),
            "total_teachers": Teacher.objects.count(),
            "total_notes": Note.objects.count(),
            "total_videos": Video.objects.count(),
            "attendance_records": Attendance.objects.count(),
            "marks_records": Marks.objects.count(),
            "total_assignments": Assignment.objects.count(),
            "total_blogs": Blog.objects.count(),
            "total_current_affairs": CurrentAffair.objects.count(),
            "recent_notices": [notice_json(row) for row in Notice.objects.order_by("-created_at")[:5]],
        }
    )


@api_view(["GET", "POST"])
def students(request):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    if request.method == "GET":
        return ok({"results": [student_json(row) for row in scoped_students(user).order_by("student_id")]})
    require_roles(request, [ROLE_ADMIN])
    data = request.data
    try:
        created_user = User(
            email=data["email"].lower().strip(),
            name=data["name"].strip(),
            role=ROLE_STUDENT,
            password_hash=hash_password(DEFAULT_PASSWORD),
            approved=True,
            first_login=True,
            force_password_change=True,
        ).save()
        row = Student(
            user=created_user,
            student_id=next_code("student", "R"),
            name=data["name"],
            email=data["email"].lower().strip(),
            class_level=str(data["class_level"]),
            division=data.get("division", ""),
            roll_number=str(data.get("roll_number", "")),
            profile_photo=data.get("profile_photo", ""),
        ).save()
    except (KeyError, NotUniqueError, ValidationError) as exc:
        return bad(str(exc))
    return ok({"student": student_json(row), "default_password": DEFAULT_PASSWORD}, status.HTTP_201_CREATED)


@api_view(["PUT", "DELETE"])
def student_detail(request, student_id):
    require_roles(request, [ROLE_ADMIN])
    row = Student.objects(id=student_id).first()
    if not row:
        return bad("Student not found", status.HTTP_404_NOT_FOUND)
    if request.method == "DELETE":
        row.user.delete()
        return ok({"message": "Student deleted"})
    data = request.data
    row.update(
        name=data.get("name", row.name),
        class_level=str(data.get("class_level", row.class_level)),
        division=data.get("division", row.division),
        roll_number=str(data.get("roll_number", row.roll_number)),
        profile_photo=data.get("profile_photo", row.profile_photo),
    )
    row.user.update(name=data.get("name", row.name), updated_at=datetime.utcnow())
    return ok({"student": student_json(Student.objects(id=student_id).first())})


@api_view(["GET", "POST"])
def teachers(request):
    require_roles(request, [ROLE_ADMIN])
    if request.method == "GET":
        return ok({"results": [teacher_json(row) for row in Teacher.objects.order_by("teacher_id")]})
    data = request.data
    try:
        created_user = User(
            email=data["email"].lower().strip(),
            name=data["name"].strip(),
            role=ROLE_TEACHER,
            password_hash=hash_password(DEFAULT_PASSWORD),
            approved=True,
            first_login=True,
            force_password_change=True,
        ).save()
        row = Teacher(
            user=created_user,
            teacher_id=next_code("teacher", "T"),
            name=data["name"],
            email=data["email"].lower().strip(),
            subjects=data.get("subjects", []),
            assigned_classes=[str(item) for item in data.get("assigned_classes", [])],
        ).save()
    except (KeyError, NotUniqueError, ValidationError) as exc:
        return bad(str(exc))
    return ok({"teacher": teacher_json(row), "default_password": DEFAULT_PASSWORD}, status.HTTP_201_CREATED)


@api_view(["PUT", "DELETE"])
def teacher_detail(request, teacher_id):
    require_roles(request, [ROLE_ADMIN])
    row = Teacher.objects(id=teacher_id).first()
    if not row:
        return bad("Teacher not found", status.HTTP_404_NOT_FOUND)
    if request.method == "DELETE":
        row.user.delete()
        return ok({"message": "Teacher deleted"})
    data = request.data
    row.update(
        name=data.get("name", row.name),
        subjects=data.get("subjects", row.subjects),
        assigned_classes=[str(item) for item in data.get("assigned_classes", row.assigned_classes)],
    )
    row.user.update(name=data.get("name", row.name), updated_at=datetime.utcnow())
    return ok({"teacher": teacher_json(Teacher.objects(id=teacher_id).first())})


@api_view(["GET", "POST", "PUT", "DELETE"])
def attendance(request):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER, ROLE_STUDENT])
    if request.method == "GET":
        if user.role == ROLE_STUDENT:
            rows = Attendance.objects(student=get_student_for_user(user)).order_by("-date")
        elif user.role == ROLE_TEACHER:
            rows = Attendance.objects(class_level__in=teacher_classes(user)).order_by("-date")
        else:
            rows = Attendance.objects.order_by("-date")
        return ok({"results": [attendance_json(row) for row in rows]})
    require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    data = request.data
    if request.method in ["PUT", "DELETE"]:
        row = Attendance.objects(id=data.get("id") or request.GET.get("id")).first()
        if not row:
            return bad("Attendance not found", status.HTTP_404_NOT_FOUND)
        enforce_teacher_class(user, row.class_level)
        if is_locked(row) and user.role != ROLE_ADMIN:
            return bad("Attendance record is locked", status.HTTP_403_FORBIDDEN)
        if request.method == "DELETE":
            require_roles(request, [ROLE_ADMIN])
            before = attendance_json(row)
            row.delete()
            log_attendance("delete", user, row, before=before)
            return ok({"message": "Attendance deleted"})
        before = attendance_json(row)
        row.update(status=data.get("status", row.status), marked_by=user, updated_at=datetime.utcnow())
        updated = Attendance.objects(id=row.id).first()
        log_attendance("update", user, updated, before=before, after=attendance_json(updated))
        notify_student(
            updated.student,
            "attendance",
            "Attendance",
            f"Your attendance for {updated.date.strftime('%d %b %Y')} was updated.",
            "/operations",
            "green",
            "attendance",
            updated.id,
        )
        return ok({"attendance": attendance_json(updated)})
    student = Student.objects(id=data.get("student")).first()
    if not student:
        return bad("Student not found", status.HTTP_404_NOT_FOUND)
    enforce_teacher_student(user, student)
    date = parse_date(data.get("date"))
    row = Attendance.objects(student=student, date=date).first()
    if row and is_locked(row) and user.role != ROLE_ADMIN:
        return bad("Attendance record is locked", status.HTTP_403_FORBIDDEN)
    now = datetime.utcnow()
    row = Attendance.objects(student=student, date=date).modify(
        upsert=True,
        new=True,
        set__class_level=student.class_level,
        set__status=data.get("status", "present"),
        set__marked_by=user,
        set__updated_at=now,
        set_on_insert__created_at=now,
    )
    log_attendance("create", user, row, after=attendance_json(row))
    notify_student(
        row.student,
        "attendance",
        "Attendance",
        f"Your attendance for {row.date.strftime('%d %b %Y')} was marked.",
        "/operations",
        "green",
        "attendance",
        row.id,
    )
    return ok({"message": "Attendance saved", "attendance": attendance_json(row)}, status.HTTP_201_CREATED)


@api_view(["POST"])
def attendance_lock(request, attendance_id):
    user = require_roles(request, [ROLE_ADMIN])
    row = Attendance.objects(id=attendance_id).first()
    if not row:
        return bad("Attendance not found", status.HTTP_404_NOT_FOUND)
    set_locked(row, user, True)
    return ok({"message": "Attendance locked", "attendance": attendance_json(row)})


@api_view(["POST"])
def attendance_unlock(request, attendance_id):
    user = require_roles(request, [ROLE_ADMIN])
    row = Attendance.objects(id=attendance_id).first()
    if not row:
        return bad("Attendance not found", status.HTTP_404_NOT_FOUND)
    set_locked(row, user, False)
    return ok({"message": "Attendance unlocked", "attendance": attendance_json(row)})


@api_view(["GET"])
def attendance_audit(request):
    require_roles(request, [ROLE_ADMIN])
    query = {}
    for key in ["class_level", "action", "student_id", "attendance_id"]:
        value = request.GET.get(key)
        if value:
            query[key] = value
    rows = list(attendance_audit_collection().find(query).sort("created_at", -1).limit(200))
    for row in rows:
        row["id"] = str(row.pop("_id"))
        for field in ["date", "created_at"]:
            row[field] = dt(row.get(field))
    return ok({"results": rows})


@api_view(["GET", "POST", "PUT", "DELETE"])
def marks(request):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER, ROLE_STUDENT])
    if request.method == "GET":
        if user.role == ROLE_STUDENT:
            rows = Marks.objects(student=get_student_for_user(user))
        elif user.role == ROLE_TEACHER:
            rows = Marks.objects(class_level__in=teacher_classes(user))
        else:
            rows = Marks.objects
        return ok({"results": [marks_json(row) for row in rows.order_by("-created_at")]})
    require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    data = request.data
    if request.method in ["PUT", "DELETE"]:
        row = Marks.objects(id=data.get("id") or request.GET.get("id")).first()
        if not row:
            return bad("Marks not found", status.HTTP_404_NOT_FOUND)
        enforce_teacher_class(user, row.class_level)
        if request.method == "DELETE":
            require_roles(request, [ROLE_ADMIN])
            row.delete()
            return ok({"message": "Marks deleted"})
        row.update(
            subject=data.get("subject", row.subject),
            exam_type=data.get("exam_type", row.exam_type),
            marks_obtained=float(data.get("marks_obtained", row.marks_obtained)),
            max_marks=float(data.get("max_marks", row.max_marks)),
            added_by=user,
        )
        updated = Marks.objects(id=row.id).first()
        notify_student(
            updated.student,
            "marks",
            "Marks",
            f"{updated.subject} {updated.exam_type} marks are available.",
            "/operations",
            "blue",
            "marks",
            updated.id,
        )
        return ok({"marks": marks_json(updated)})
    student = Student.objects(id=data.get("student")).first()
    if not student:
        return bad("Student not found", status.HTTP_404_NOT_FOUND)
    enforce_teacher_student(user, student)
    row = Marks(
        student=student,
        class_level=student.class_level,
        subject=data.get("subject"),
        exam_type=data.get("exam_type"),
        marks_obtained=float(data.get("marks_obtained")),
        max_marks=float(data.get("max_marks")),
        added_by=user,
    ).save()
    notify_student(
        row.student,
        "marks",
        "Marks",
        f"{row.subject} {row.exam_type} marks are available.",
        "/operations",
        "blue",
        "marks",
        row.id,
    )
    return ok({"marks": marks_json(row)}, status.HTTP_201_CREATED)


def class_query(request, model):
    user = current_user(request)
    class_level = request.GET.get("class_level")
    query = model.objects
    if user.role == ROLE_STUDENT:
        student = get_student_for_user(user)
        query = query(class_level__in=[student.class_level, "all"] if model == Notice else [student.class_level])
    elif user.role == ROLE_TEACHER:
        allowed = teacher_classes(user) + (["all"] if model == Notice else [])
        query = query(class_level__in=allowed)
        if class_level:
            query = query(class_level=class_level) if class_level in allowed else query(class_level="__none__")
    elif class_level:
        query = query(class_level=class_level)
    return user, query


def content_view(model, fields, owner_field):
    @api_view(["GET", "POST", "PUT", "DELETE"])
    def handler(request):
        user, query = class_query(request, model)
        if request.method == "GET":
            rows = list(query.order_by("-created_at"))
            results = [simple_json(row, fields) for row in rows]
            if model == Note and user.role == ROLE_STUDENT:
                student = get_student_for_user(user)
                bookmarked_ids = {str(row.note.id) for row in NoteBookmark.objects(student=student, note__in=rows)} if student else set()
                for item in results:
                    item["bookmarked"] = item["id"] in bookmarked_ids
            return ok({"results": results})
        require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
        data = request.data
        if request.method in ["PUT", "DELETE"]:
            row = model.objects(id=data.get("id") or request.GET.get("id")).first()
            if not row:
                return bad("Item not found", status.HTTP_404_NOT_FOUND)
            if hasattr(row, "class_level"):
                enforce_teacher_class(user, row.class_level)
            enforce_owner(user, row, owner_field)
            if request.method == "DELETE":
                row.delete()
                return ok({"message": "Item deleted"})
            payload = {field: data.get(field, getattr(row, field)) for field in fields if field != "created_at"}
            if "class_level" in payload:
                payload["class_level"] = str(payload["class_level"])
                enforce_teacher_class(user, payload["class_level"])
            row.update(**{f"set__{key}": value for key, value in payload.items()})
            return ok({"item": simple_json(model.objects(id=row.id).first(), fields)})
        payload = {field: data.get(field) for field in fields if field != "created_at"}
        if "class_level" in payload:
            payload["class_level"] = str(payload["class_level"])
            enforce_teacher_class(user, payload["class_level"])
        payload[owner_field] = user
        row = model(**payload).save()
        if model == Video:
            notify_students_for_class(
                row.class_level,
                "video",
                "Learning Content",
                f"New video lecture added: {row.title}",
                "/learning",
                "blue",
                "learning",
                row.id,
            )
        elif model == Note:
            notify_students_for_class(
                row.class_level,
                "note",
                "Learning Content",
                f"New study note added: {row.title}",
                "/learning",
                "blue",
                "learning",
                row.id,
            )
        return ok({"item": simple_json(row, fields)}, status.HTTP_201_CREATED)
    return handler


videos = content_view(Video, ["title", "class_level", "subject", "chapter", "description", "youtube_url", "created_at"], "uploaded_by")
notes = content_view(Note, ["title", "class_level", "subject", "chapter", "pdf_url", "created_at"], "uploaded_by")


@api_view(["GET", "POST", "PUT", "DELETE"])
def notices(request):
    user, query = class_query(request, Notice)
    if request.method == "GET":
        return ok({"results": [notice_json(row) for row in query.order_by("-created_at")]})
    require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    data = request.data
    if request.method in ["PUT", "DELETE"]:
        row = Notice.objects(id=data.get("id") or request.GET.get("id")).first()
        if not row:
            return bad("Notice not found", status.HTTP_404_NOT_FOUND)
        enforce_teacher_class(user, row.class_level)
        enforce_owner(user, row, "created_by")
        if request.method == "DELETE":
            row.delete()
            return ok({"message": "Notice deleted"})
        class_level = str(data.get("class_level", row.class_level))
        enforce_teacher_class(user, class_level)
        row.update(title=data.get("title", row.title), body=data.get("body", row.body), class_level=class_level, updated_at=datetime.utcnow())
        return ok({"notice": notice_json(Notice.objects(id=row.id).first())})
    class_level = str(data.get("class_level", "all"))
    enforce_teacher_class(user, class_level)
    row = Notice(title=data.get("title"), body=data.get("body"), class_level=class_level, created_by=user).save()
    notify_students_for_class(
        row.class_level,
        "notice",
        "Notice",
        row.title,
        "/learning",
        "red",
        "notice",
        row.id,
    )
    return ok({"notice": notice_json(row)}, status.HTTP_201_CREATED)


@api_view(["GET", "POST", "DELETE"])
def timetables(request):
    user = current_user(request)
    if request.method == "GET":
        class_level = request.GET.get("class_level")
        if user.role == ROLE_STUDENT:
            class_level = get_student_for_user(user).class_level
        if user.role == ROLE_TEACHER:
            assigned = teacher_classes(user)
            rows = Timetable.objects(class_level=class_level) if class_level in assigned else Timetable.objects(class_level__in=assigned)
        else:
            rows = Timetable.objects(class_level=class_level) if class_level else Timetable.objects
        return ok({"results": [timetable_json(row) for row in rows]})
    require_roles(request, [ROLE_ADMIN])
    if request.method == "DELETE":
        Timetable.objects(id=request.data.get("id") or request.GET.get("id")).delete()
        return ok({"message": "Timetable deleted"})
    data = request.data
    periods = [TimetablePeriod(**period) for period in data.get("periods", [])]
    row = Timetable.objects(class_level=str(data.get("class_level"))).modify(upsert=True, new=True, set__periods=periods, set__updated_at=datetime.utcnow())
    notify_students_for_class(
        row.class_level,
        "timetable",
        "Timetable",
        "Your class timetable has been updated.",
        "/operations",
        "amber",
        "timetable",
        row.id,
    )
    return ok({"timetable": timetable_json(row)})


PAYMENT_MODES = ["Cash", "UPI", "Bank Transfer", "Cheque", "Other"]


def fee_structure_json(row):
    return {
        **simple_json(row, ["class_level", "annual_fee", "installments", "updated_at"]),
        "due_date": dt(getattr(row, "due_date", None)),
    }


def payment_json(row):
    return {
        "id": oid(row),
        "student_id": row.student.student_id if row.student else "",
        "student_name": row.student.name if row.student else "",
        "class_level": row.class_level,
        "amount": row.amount,
        "payment_date": dt(row.payment_date),
        "payment_mode": row.payment_mode,
        "reference": row.reference,
        "installment": row.installment,
        "note": row.note,
        "recorded_by": user_json(row.recorded_by),
        "created_at": dt(row.created_at),
    }


def fee_status(total_fee, paid_amount, due_date):
    pending = max(0, total_fee - paid_amount)
    if pending <= 0:
        return "Paid"
    if due_date and datetime.utcnow().date() > due_date.date():
        return "Overdue"
    if paid_amount > 0:
        return "Partial"
    return "Pending"


def fee_scope_for_user(user, requested_class=None):
    if user.role == ROLE_STUDENT:
        student = get_student_for_user(user)
        return [student.class_level] if student else []
    if user.role == ROLE_TEACHER:
        assigned = teacher_classes(user)
        return [requested_class] if requested_class in assigned else assigned
    return [requested_class] if requested_class else None


def fee_student_rows(user, structures, requested_status=""):
    structure_map = {row.class_level: row for row in structures}
    query = scoped_students(user)
    if structure_map:
        query = query(class_level__in=list(structure_map.keys()))
    rows = []
    summary = {"total_fees": 0, "paid": 0, "pending": 0, "overdue": 0}
    for student in query.order_by("student_id"):
        fee = structure_map.get(student.class_level)
        if not fee:
            continue
        payments = list(FeePayment.objects(student=student).order_by("-payment_date"))
        paid = sum(float(payment.amount or 0) for payment in payments)
        total = float(fee.annual_fee or 0)
        pending = max(0, total - paid)
        status_value = fee_status(total, paid, getattr(fee, "due_date", None))
        if requested_status and status_value.lower() != requested_status.lower():
            continue
        summary["total_fees"] += total
        summary["paid"] += paid
        summary["pending"] += pending
        if status_value == "Overdue":
            summary["overdue"] += pending
        rows.append(
            {
                "student": student_json(student),
                "class_level": student.class_level,
                "total_fee": total,
                "paid": paid,
                "pending": pending,
                "due_date": dt(getattr(fee, "due_date", None)),
                "status": status_value,
                "installments": fee.installments,
                "payments": [payment_json(payment) for payment in payments],
            }
        )
    return rows, summary


def parse_installments(value):
    if isinstance(value, dict):
        return value
    return {}


@api_view(["GET", "POST", "DELETE"])
def fees(request):
    user = current_user(request)
    if request.method == "GET":
        requested_class = request.GET.get("class_level")
        status_filter = request.GET.get("status", "")
        scope = fee_scope_for_user(user, requested_class)
        rows = Fee.objects(class_level__in=scope) if scope is not None else Fee.objects
        structures = list(rows.order_by("class_level"))
        student_rows, summary = fee_student_rows(user, structures, status_filter)
        return ok(
            {
                "results": [fee_structure_json(row) for row in structures],
                "student_records": student_rows,
                "summary": summary,
                "payment_modes": PAYMENT_MODES,
            }
        )
    require_roles(request, [ROLE_ADMIN])
    if request.method == "DELETE":
        row = Fee.objects(id=request.data.get("id") or request.GET.get("id")).first()
        if not row:
            return bad("Fee structure not found", status.HTTP_404_NOT_FOUND)
        if FeePayment.objects(class_level=row.class_level).first():
            return bad("Cannot delete fee structure with payment history.")
        row.delete()
        return ok({"message": "Fee structure deleted"})
    data = request.data
    class_level = str(data.get("class_level", ""))
    if class_level not in [str(item) for item in range(6, 13)]:
        return bad("Select a valid class.")
    try:
        annual_fee = float(data.get("annual_fee"))
    except (TypeError, ValueError):
        return bad("Annual fee must be a valid amount.")
    if annual_fee <= 0:
        return bad("Annual fee must be greater than 0.")
    due_date = parse_date(data.get("due_date")) if data.get("due_date") else None
    installments = parse_installments(data.get("installments", {}))
    installment_total = sum(float(value or 0) for value in installments.values()) if installments else 0
    if installments and abs(installment_total - annual_fee) > 0.01:
        return bad("Installment totals must match annual fee.")
    row = Fee.objects(class_level=class_level).modify(
        upsert=True,
        new=True,
        set__annual_fee=annual_fee,
        set__installments=installments,
        set__due_date=due_date,
        set__updated_at=datetime.utcnow(),
    )
    notify_students_for_class(row.class_level, "fee", "Fee", "Fee structure has been updated.", "/operations", "amber", "fee", row.id)
    return ok({"fee": fee_structure_json(row)})


@api_view(["POST"])
def fee_payments(request):
    user = require_roles(request, [ROLE_ADMIN])
    data = request.data
    student = Student.objects(id=data.get("student") or data.get("student_id")).first()
    if not student:
        return bad("Student not found", status.HTTP_404_NOT_FOUND)
    fee = Fee.objects(class_level=student.class_level).first()
    if not fee:
        return bad("Create a fee structure for this student's class first.")
    try:
        amount = float(data.get("amount"))
    except (TypeError, ValueError):
        return bad("Payment amount must be valid.")
    if amount <= 0:
        return bad("Payment amount must be greater than 0.")
    paid = sum(float(payment.amount or 0) for payment in FeePayment.objects(student=student))
    if paid + amount > float(fee.annual_fee or 0):
        return bad("Payment amount cannot exceed pending fee.")
    mode = data.get("payment_mode", "Cash")
    if mode not in PAYMENT_MODES:
        return bad("Select a valid payment mode.")
    row = FeePayment(
        student=student,
        class_level=student.class_level,
        amount=amount,
        payment_date=parse_date(data.get("payment_date")) if data.get("payment_date") else datetime.utcnow(),
        payment_mode=mode,
        reference=data.get("reference", "").strip(),
        installment=data.get("installment", "").strip(),
        note=data.get("note", "").strip(),
        recorded_by=user,
    ).save()
    notify_student(student, "fee_payment", "Fee Payment", f"Payment of Rs {amount:g} has been recorded.", "/operations", "green", "fee", row.id)
    return ok({"payment": payment_json(row)}, status.HTTP_201_CREATED)


@api_view(["GET", "POST", "PUT", "DELETE"])
def blogs(request):
    user = current_user(request)
    if request.method == "GET":
        return ok({"results": [simple_json(row, ["title", "category", "content", "published", "created_at"]) for row in Blog.objects(published=True).order_by("-created_at")]})
    require_roles(request, [ROLE_ADMIN])
    data = request.data
    if request.method in ["PUT", "DELETE"]:
        row = Blog.objects(id=data.get("id") or request.GET.get("id")).first()
        if not row:
            return bad("Blog not found", status.HTTP_404_NOT_FOUND)
        if request.method == "DELETE":
            row.delete()
            return ok({"message": "Blog deleted"})
        row.update(title=data.get("title", row.title), category=data.get("category", row.category), content=data.get("content", row.content), published=data.get("published", row.published))
        return ok({"item": simple_json(Blog.objects(id=row.id).first(), ["title", "category", "content", "published", "created_at"])})
    row = Blog(title=data.get("title"), category=data.get("category"), content=data.get("content"), published=data.get("published", True), author=user).save()
    if row.published:
        notify_all_students(
            "blog",
            "Blog",
            f"New study article published: {row.title}",
            "/learning",
            "green",
            "blog",
            row.id,
        )
    return ok({"item": simple_json(row, ["title", "category", "content", "published", "created_at"])}, status.HTTP_201_CREATED)


@api_view(["GET", "POST", "PUT", "DELETE"])
def current_affairs(request):
    user = current_user(request)
    fields = ["title", "summary", "content", "category", "source_url", "source_name", "image_url", "generated_by_ai", "digest_date", "published_on"]
    if request.method == "GET":
        return ok({"results": [simple_json(row, fields) for row in CurrentAffair.objects.order_by("-published_on")]})
    require_roles(request, [ROLE_ADMIN])
    data = request.data
    if request.method in ["PUT", "DELETE"]:
        row = CurrentAffair.objects(id=data.get("id") or request.GET.get("id")).first()
        if not row:
            return bad("Current affair not found", status.HTTP_404_NOT_FOUND)
        if request.method == "DELETE":
            row.delete()
            return ok({"message": "Current affair deleted"})
        row.update(
            title=data.get("title", row.title),
            summary=data.get("summary", row.summary),
            content=data.get("content", row.content),
            category=data.get("category", row.category),
            source_url=data.get("source_url", row.source_url),
            source_name=data.get("source_name", row.source_name),
            image_url=data.get("image_url", row.image_url),
        )
        return ok({"item": simple_json(CurrentAffair.objects(id=row.id).first(), fields)})
    row = CurrentAffair(
        title=data.get("title"),
        summary=data.get("summary"),
        content=data.get("content", ""),
        category=data.get("category", "Educational News"),
        source_url=data.get("source_url", ""),
        source_name=data.get("source_name", ""),
        image_url=data.get("image_url", ""),
        created_by=user,
    ).save()
    notify_all_students(
        "current_affair",
        "Current Affairs",
        row.title,
        "/learning",
        "violet",
        "current_affairs",
        row.id,
    )
    return ok({"item": simple_json(row, fields)}, status.HTTP_201_CREATED)


@api_view(["GET", "POST", "PUT", "DELETE"])
def assignments(request):
    user, query = class_query(request, Assignment)
    if request.method == "GET":
        return ok({"results": [assignment_json(row, user) for row in query.order_by("deadline")]})
    require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    data = request.data
    if request.method in ["PUT", "DELETE"]:
        row = Assignment.objects(id=data.get("id") or request.GET.get("id")).first()
        if not row:
            return bad("Assignment not found", status.HTTP_404_NOT_FOUND)
        enforce_teacher_class(user, row.class_level)
        enforce_teacher_subject(user, row.subject)
        enforce_owner(user, row, "created_by")
        if request.method == "DELETE":
            row.delete()
            return ok({"message": "Assignment deleted"})
        try:
            row.update(**assignment_payload(data, user, row))
        except ValueError as exc:
            return bad(str(exc))
        return ok({"assignment": assignment_json(Assignment.objects(id=row.id).first(), user)})
    try:
        payload = assignment_payload(data, user)
    except ValueError as exc:
        return bad(str(exc))
    row = Assignment(**payload, created_by=user).save()
    notify_students_for_class(
        row.class_level,
        "assignment",
        "Assignment",
        f"New {row.subject} assignment added: {row.title}",
        "/operations",
        "red",
        "assignment",
        row.id,
    )
    return ok({"assignment": assignment_json(row, user)}, status.HTTP_201_CREATED)


@api_view(["POST"])
def assignment_submit(request, assignment_id):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    row = Assignment.objects(id=assignment_id).first()
    if not row:
        return bad("Assignment not found", status.HTTP_404_NOT_FOUND)
    if not student or student.class_level != row.class_level:
        raise PermissionDenied("Students can only submit assignments for their class")
    now = datetime.utcnow()
    submission = AssignmentSubmission.objects(assignment=row, student=student).first()
    fields = {
        "answer_text": str(request.data.get("answer_text", "") or "").strip(),
        "file_url": str(request.data.get("file_url", "") or "").strip(),
        "status": "late" if row.deadline and row.deadline < now else "submitted",
        "submitted_at": now,
        "updated_at": now,
    }
    if submission:
        submission.update(**fields)
        submission = AssignmentSubmission.objects(id=submission.id).first()
    else:
        submission = AssignmentSubmission(assignment=row, student=student, **fields).save()
    return ok({"submission": assignment_submission_json(submission)})


def normalize_answer(value):
    return " ".join(str(value or "").strip().lower().split())


def objective_correct_answer(question):
    if question.question_type == "mcq":
        answer = str(question.correct_answer or "").strip()
        option_index = {"a": 0, "b": 1, "c": 2, "d": 3}.get(answer.lower())
        if option_index is not None and option_index < len(question.options):
            return question.options[option_index]
    return question.correct_answer


def today_key():
    return date.today().isoformat()


def practice_count():
    try:
        return max(1, min(int(os.getenv("DAILY_PRACTICE_QUESTION_COUNT", "10")), 25))
    except (TypeError, ValueError):
        return 10


def question_bank_json(question, include_correct=False):
    data = {
        "id": oid(question),
        "class_level": question.class_level,
        "subject": question.subject,
        "chapter": question.chapter,
        "question_type": question.question_type,
        "difficulty": question.difficulty,
        "text": question.text,
        "options": question.options,
        "marks": question.marks,
        "explanation": question.explanation if include_correct else "",
        "created_by": user_json(question.created_by),
        "created_at": dt(question.created_at),
        "updated_at": dt(question.updated_at),
    }
    if include_correct:
        data["correct_answer"] = question.correct_answer
        data["expected_answer"] = question.expected_answer
    return data


def question_bank_from_data(data, user, row=None):
    class_level = str(data.get("class_level", row.class_level if row else ""))
    subject = str(data.get("subject", row.subject if row else "")).strip()
    enforce_teacher_question_scope(user, class_level, subject)
    qtype = data.get("question_type", row.question_type if row else "mcq")
    if qtype not in ["mcq", "true_false", "short", "long"]:
        raise ValueError("Invalid question type")
    difficulty = data.get("difficulty", row.difficulty if row else "Medium")
    if difficulty not in ["Easy", "Medium", "Hard"]:
        raise ValueError("Invalid difficulty")
    options = data.get("options", row.options if row else [])
    if qtype == "mcq":
        options = [str(item).strip() for item in options if str(item).strip()][:4]
        if len(options) < 2:
            raise ValueError("MCQ questions need at least two options")
    elif qtype == "true_false":
        options = ["True", "False"]
    else:
        options = []
    return {
        "class_level": class_level,
        "subject": subject,
        "chapter": str(data.get("chapter", row.chapter if row else "")).strip(),
        "question_type": qtype,
        "difficulty": difficulty,
        "text": str(data.get("text", row.text if row else "")).strip(),
        "options": options,
        "correct_answer": str(data.get("correct_answer", row.correct_answer if row else "")).strip(),
        "expected_answer": str(data.get("expected_answer", row.expected_answer if row else "")).strip(),
        "explanation": str(data.get("explanation", row.explanation if row else "")).strip(),
        "marks": float(data.get("marks", row.marks if row else 1) or 1),
    }


def exam_question_from_bank(question, order):
    return ExamQuestion(
        question_id=uuid4().hex,
        question_bank_id=oid(question),
        text=question.text,
        question_type=question.question_type,
        marks=question.marks,
        options=question.options,
        correct_answer=question.correct_answer,
        expected_answer=question.expected_answer,
        explanation=question.explanation,
        chapter=question.chapter,
        difficulty=question.difficulty,
        order=order,
    )


def mistake_json(row):
    question = row.question
    correct_answer = row.correct_answer or (objective_correct_answer(question) if question and question.question_type in ["mcq", "true_false"] else getattr(question, "expected_answer", "")) or ""
    explanation = row.explanation or (question.explanation if question else "")
    return {
        "id": oid(row),
        "question": question_bank_json(question, include_correct=True) if question else None,
        "source": row.source,
        "subject": row.subject,
        "chapter": row.chapter,
        "student_answer": row.last_answer,
        "correct_answer": correct_answer,
        "explanation": explanation,
        "wrong_attempts": row.wrong_attempts,
        "retry_attempts": getattr(row, "retry_attempts", 0),
        "correct_streak": row.correct_streak,
        "resolved": row.resolved,
        "status": "Corrected" if row.resolved else "Needs Practice",
        "last_wrong_at": dt(row.last_wrong_at),
        "last_retry_at": dt(getattr(row, "last_retry_at", None)),
        "last_retry_answer": getattr(row, "last_retry_answer", ""),
        "last_retry_correct": getattr(row, "last_retry_correct", False),
        "corrected_at": dt(getattr(row, "corrected_at", None)),
        "updated_at": dt(row.updated_at),
    }


def topic_status(attempts, correct):
    if attempts < 3:
        return "Needs Practice"
    accuracy = (correct / attempts) * 100 if attempts else 0
    if accuracy >= 75:
        return "Strong"
    if accuracy >= 50:
        return "Needs Practice"
    return "Weak"


def update_practice_tracking(student, question, correct, source="practice", answer_text=""):
    if not question:
        return
    now = datetime.utcnow()
    correct_answer = objective_correct_answer(question) if question.question_type in ["mcq", "true_false"] else question.expected_answer
    performance = TopicPerformance.objects(student=student, subject=question.subject, chapter=question.chapter).first()
    if performance:
        attempts = performance.attempts + 1
        correct_count = performance.correct + (1 if correct else 0)
        performance.update(
            attempts=attempts,
            correct=correct_count,
            accuracy=round((correct_count / attempts) * 100, 2),
            status=topic_status(attempts, correct_count),
            updated_at=now,
        )
    else:
        TopicPerformance(
            student=student,
            class_level=student.class_level,
            subject=question.subject,
            chapter=question.chapter,
            attempts=1,
            correct=1 if correct else 0,
            accuracy=100 if correct else 0,
            status=topic_status(1, 1 if correct else 0),
            updated_at=now,
        ).save()
    mistake = StudentMistake.objects(student=student, question=question).first()
    if correct:
        if mistake:
            streak = mistake.correct_streak + 1
            mistake.update(correct_streak=streak, resolved=streak >= 2, updated_at=now, corrected_at=now if streak >= 2 else mistake.corrected_at)
        return
    if mistake:
        mistake.update(
            source=source,
            last_answer=answer_text,
            correct_answer=correct_answer or "",
            explanation=question.explanation or "",
            wrong_attempts=mistake.wrong_attempts + 1,
            correct_streak=0,
            resolved=False,
            last_wrong_at=now,
            updated_at=now,
        )
    else:
        StudentMistake(
            student=student,
            question=question,
            source=source,
            subject=question.subject,
            chapter=question.chapter,
            last_answer=answer_text,
            correct_answer=correct_answer or "",
            explanation=question.explanation or "",
            wrong_attempts=1,
            correct_streak=0,
            resolved=False,
            last_wrong_at=now,
            updated_at=now,
        ).save()


def practice_session_json(session, include_correct=False):
    answers = {oid(answer.question): answer for answer in session.answers if answer.question}
    questions = []
    for question in session.questions:
        item = question_bank_json(question, include_correct=include_correct)
        answer = answers.get(oid(question))
        item["answer"] = answer.answer if answer else ""
        item["correct"] = answer.correct if answer and include_correct else False
        item["marks_awarded"] = answer.marks_awarded if answer and include_correct else 0
        questions.append(item)
    return {
        "id": oid(session),
        "session_type": session.session_type,
        "session_date": session.session_date,
        "class_level": session.class_level,
        "subject": session.subject,
        "status": session.status,
        "questions_per_day": practice_count(),
        "total_questions": session.total_questions,
        "correct_count": session.correct_count,
        "incorrect_count": session.incorrect_count,
        "score": session.score,
        "accuracy": session.accuracy,
        "topic_performance": session.topic_performance,
        "started_at": dt(session.started_at),
        "submitted_at": dt(session.submitted_at),
        "questions": questions,
    }


def select_daily_questions(student, limit):
    weak_topics = list(TopicPerformance.objects(student=student, status__in=["Weak", "Needs Practice"]).order_by("accuracy")[:8])
    selected = []
    selected_ids = set()
    for topic in weak_topics:
        for question in QuestionBankQuestion.objects(class_level=student.class_level, subject=topic.subject, chapter=topic.chapter).order_by("-created_at")[:limit]:
            if oid(question) not in selected_ids:
                selected.append(question)
                selected_ids.add(oid(question))
            if len(selected) >= limit:
                return selected
    for question in QuestionBankQuestion.objects(class_level=student.class_level).order_by("difficulty", "-created_at"):
        if oid(question) not in selected_ids:
            selected.append(question)
            selected_ids.add(oid(question))
        if len(selected) >= limit:
            break
    return selected


def get_or_create_daily_session(student):
    key = today_key()
    existing = PracticeSession.objects(student=student, session_type="daily", session_date=key).first()
    if existing:
        if existing.status == "in_progress" and not existing.questions:
            questions = select_daily_questions(student, practice_count())
            if questions:
                existing.update(questions=questions, total_questions=len(questions), updated_at=datetime.utcnow())
                return PracticeSession.objects(id=existing.id).first()
        return existing
    questions = select_daily_questions(student, practice_count())
    return PracticeSession(
        student=student,
        session_type="daily",
        session_date=key,
        class_level=student.class_level,
        questions=questions,
        total_questions=len(questions),
    ).save()


def submit_practice_session(session):
    if session.status == "submitted":
        return session
    by_id = {oid(answer.question): answer for answer in session.answers if answer.question}
    answers = []
    correct_count = 0
    score = 0
    topic_map = {}
    now = datetime.utcnow()
    for question in session.questions:
        answer = by_id.get(oid(question)) or PracticeAnswer(question=question)
        if question.question_type in ["mcq", "true_false"]:
            is_correct = normalize_answer(answer.answer) == normalize_answer(objective_correct_answer(question))
        else:
            is_correct = bool(question.expected_answer) and normalize_answer(answer.answer) == normalize_answer(question.expected_answer)
        answer.correct = is_correct
        answer.marks_awarded = question.marks if is_correct else 0
        answer.answered_at = answer.answered_at or now
        answers.append(answer)
        correct_count += 1 if is_correct else 0
        score += answer.marks_awarded
        key = (question.subject, question.chapter)
        current = topic_map.get(key) or {"subject": question.subject, "chapter": question.chapter, "attempts": 0, "correct": 0}
        current["attempts"] += 1
        current["correct"] += 1 if is_correct else 0
        topic_map[key] = current
        update_practice_tracking(session.student, question, is_correct, "practice", answer.answer)
    topic_rows = []
    for row in topic_map.values():
        accuracy = round((row["correct"] / row["attempts"]) * 100, 2) if row["attempts"] else 0
        topic_rows.append({**row, "accuracy": accuracy, "status": topic_status(row["attempts"], row["correct"])})
    total = len(session.questions)
    session.update(
        answers=answers,
        status="submitted",
        total_questions=total,
        correct_count=correct_count,
        incorrect_count=max(0, total - correct_count),
        score=score,
        accuracy=round((correct_count / total) * 100, 2) if total else 0,
        topic_performance=topic_rows,
        submitted_at=now,
        updated_at=now,
    )
    return PracticeSession.objects(id=session.id).first()


def evaluate_practice_answer(question, answer):
    if question.question_type in ["mcq", "true_false"]:
        return normalize_answer(answer) == normalize_answer(objective_correct_answer(question))
    return bool(question.expected_answer) and normalize_answer(answer) == normalize_answer(question.expected_answer)


def topic_performance_json(row):
    return {
        "id": oid(row),
        "student": student_json(row.student) if row.student else None,
        "class_level": row.class_level,
        "subject": row.subject,
        "chapter": row.chapter,
        "attempts": row.attempts,
        "correct": row.correct,
        "accuracy": row.accuracy,
        "status": row.status,
        "updated_at": dt(row.updated_at),
    }


def learning_links_for(student, subject, chapter):
    notes = Note.objects(class_level=student.class_level, subject=subject, chapter=chapter).first()
    videos = Video.objects(class_level=student.class_level, subject=subject, chapter=chapter).first()
    if notes or videos:
        return "/learning"
    return ""


def study_task(source_key, title, category, subject="", chapter="", reason="", minutes=20, link="", scope="today"):
    return StudyPlanTask(
        task_id=uuid4().hex,
        source_key=source_key,
        title=title,
        category=category,
        subject=subject,
        chapter=chapter,
        reason=reason,
        minutes=minutes,
        link=link,
        scope=scope,
        status="Pending",
    )


def planner_candidates(student):
    now = datetime.utcnow()
    week_end = now + timedelta(days=7)
    tasks = []

    exams = Exam.objects(class_level=student.class_level, is_published=True, start_time__gte=now, start_time__lte=week_end).order_by("start_time")[:4]
    for exam in exams:
        attempt = ExamAttempt.objects(exam=exam, student=student).first()
        if attempt and attempt.submitted_at:
            continue
        scope = "today" if exam.start_time.date() == now.date() else "week"
        tasks.append(study_task(
            f"exam:{oid(exam)}",
            f"Prepare for {exam.name}",
            "Exam",
            exam.subject,
            "",
            f"Upcoming exam on {exam.start_time.strftime('%d %b')}",
            35,
            "/operations",
            scope,
        ))

    assignments = Assignment.objects(class_level=student.class_level, deadline__gte=now, deadline__lte=week_end).order_by("deadline")[:5]
    for assignment in assignments:
        if AssignmentSubmission.objects(assignment=assignment, student=student).first():
            continue
        scope = "today" if assignment.deadline.date() == now.date() else "week"
        tasks.append(study_task(
            f"assignment:{oid(assignment)}",
            f"Complete assignment: {assignment.title}",
            "Assignment",
            assignment.subject,
            "",
            f"Due {assignment.deadline.strftime('%d %b')}",
            25,
            "/operations",
            scope,
        ))

    weak_rows = TopicPerformance.objects(student=student, status__in=["Weak", "Needs Practice"]).order_by("accuracy", "-attempts")[:4]
    for row in weak_rows:
        tasks.append(study_task(
            f"weak:{row.subject}:{row.chapter}",
            f"Revise {row.chapter}",
            "Weak Topic",
            row.subject,
            row.chapter,
            f"{row.status} topic with {round(row.accuracy or 0)}% accuracy",
            30 if row.status == "Weak" else 20,
            learning_links_for(student, row.subject, row.chapter) or "/practice-progress",
            "today",
        ))

    mistakes = StudentMistake.objects(student=student, resolved=False).order_by("-wrong_attempts", "-last_wrong_at")[:4]
    for mistake in mistakes:
        tasks.append(study_task(
            f"mistake:{oid(mistake)}",
            f"Retry mistake: {mistake.chapter}",
            "My Mistakes",
            mistake.subject,
            mistake.chapter,
            f"{mistake.wrong_attempts} wrong attempt(s) recorded",
            20,
            "/practice-progress",
            "today",
        ))

    bookmarked_note_ids = {oid(bookmark.note) for bookmark in NoteBookmark.objects(student=student)}
    notes = Note.objects(class_level=student.class_level).order_by("-created_at")[:8]
    for note in notes:
        if oid(note) in bookmarked_note_ids:
            continue
        tasks.append(study_task(
            f"note:{oid(note)}",
            f"Review notes: {note.title}",
            "Learning",
            note.subject,
            note.chapter,
            "Recent class note not bookmarked yet",
            20,
            "/learning",
            "week",
        ))
        if len([task for task in tasks if task.category == "Learning"]) >= 2:
            break

    if not tasks:
        tasks.append(study_task(
            f"revision:{student.class_level}:{today_key()}",
            "General class revision",
            "Revision",
            "",
            "",
            "No urgent weak topic, assignment, or exam found",
            20,
            "/learning",
            "today",
        ))
    return tasks


def merge_plan_tasks(plan, candidates):
    existing_by_source = {task.source_key: task for task in plan.tasks if getattr(task, "source_key", "")}
    existing_sources = set(existing_by_source)
    merged = list(plan.tasks)
    for candidate in candidates:
        if candidate.source_key in existing_sources:
            current = existing_by_source[candidate.source_key]
            current.title = candidate.title
            current.category = candidate.category
            current.subject = candidate.subject
            current.chapter = candidate.chapter
            current.reason = candidate.reason
            current.minutes = candidate.minutes
            current.link = candidate.link
            current.scope = candidate.scope
            continue
        merged.append(candidate)
        existing_sources.add(candidate.source_key)
    return merged[:10]


def get_or_create_study_plan(student):
    key = today_key()
    plan = StudyPlan.objects(student=student, plan_date=key).first()
    if plan:
        tasks = merge_plan_tasks(plan, planner_candidates(student))
        plan.update(tasks=tasks, updated_at=datetime.utcnow())
        return StudyPlan.objects(id=plan.id).first()
    tasks = planner_candidates(student)
    return StudyPlan(student=student, plan_date=key, tasks=tasks[:10]).save()


def study_plan_json(plan):
    return {
        "id": oid(plan),
        "plan_date": plan.plan_date,
        "created_at": dt(plan.created_at),
        "updated_at": dt(plan.updated_at),
        "tasks": [
            {
                "task_id": task.task_id,
                "title": task.title,
                "category": task.category,
                "subject": task.subject,
                "chapter": task.chapter,
                "reason": getattr(task, "reason", ""),
                "minutes": task.minutes,
                "link": task.link,
                "source_key": getattr(task, "source_key", ""),
                "scope": getattr(task, "scope", "today"),
                "status": getattr(task, "status", "Completed" if task.completed else "Pending"),
                "completed": task.completed,
                "completed_at": dt(task.completed_at),
            }
            for task in plan.tasks
        ],
        "today": [
            task.task_id for task in plan.tasks
            if getattr(task, "scope", "today") == "today"
        ],
        "week": [
            task.task_id for task in plan.tasks
            if getattr(task, "scope", "today") == "week"
        ],
    }


def process_published_exam_attempts(exam):
    bank_questions = {question.question_id: QuestionBankQuestion.objects(id=question.question_bank_id).first() for question in exam.questions if question.question_bank_id}
    for attempt in ExamAttempt.objects(exam=exam):
        answers = {answer.question_id: answer for answer in attempt.answers}
        for question in exam.questions:
            bank_question = bank_questions.get(question.question_id)
            if not bank_question or question.question_type not in ["mcq", "true_false"]:
                continue
            answer = answers.get(question.question_id)
            correct = bool(answer and normalize_answer(answer.answer) == normalize_answer(objective_correct_answer(question)))
            update_practice_tracking(attempt.student, bank_question, correct, "exam", answer.answer if answer else "")


QUESTION_BULK_HEADERS = ["Class", "Subject", "Chapter", "Type", "Difficulty", "Marks", "Question", "Option A", "Option B", "Option C", "Option D", "Correct Answer", "Explanation"]


@api_view(["POST"])
def question_bank_bulk_import(request):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    upload = request.FILES.get("file")
    if not upload:
        return bad("Upload a CSV file with columns: " + ", ".join(QUESTION_BULK_HEADERS))
    try:
        text = upload.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        return bad("Unable to read this file. Please upload a CSV file.")
    reader = csv.DictReader(io.StringIO(text))
    created, errors = [], []
    for index, row in enumerate(reader, start=2):
        try:
            qtype_raw = str(row.get("Type") or row.get("type") or "mcq").strip().lower()
            qtype = {"mcq": "mcq", "true/false": "true_false", "true_false": "true_false", "short": "short", "long": "long"}.get(qtype_raw, qtype_raw)
            options = [row.get("Option A") or "", row.get("Option B") or "", row.get("Option C") or "", row.get("Option D") or ""]
            body = {
                "class_level": row.get("Class") or row.get("class"),
                "subject": row.get("Subject") or row.get("subject"),
                "chapter": row.get("Chapter") or row.get("chapter") or "",
                "question_type": qtype,
                "difficulty": (row.get("Difficulty") or row.get("difficulty") or "Medium").strip() or "Medium",
                "marks": row.get("Marks") or row.get("marks") or 1,
                "text": row.get("Question") or row.get("question"),
                "options": options,
                "correct_answer": row.get("Correct Answer") or row.get("correct_answer") or "",
                "expected_answer": row.get("Correct Answer") or row.get("correct_answer") or "" if qtype in ("short", "long") else "",
                "explanation": row.get("Explanation") or row.get("explanation") or "",
            }
            if not body["text"] or not str(body["text"]).strip():
                raise ValueError("Question text is required")
            payload = question_bank_from_data(body, user)
            question = QuestionBankQuestion(**payload, created_by=user).save()
            created.append(question_bank_json(question, include_correct=True))
        except PermissionDenied as exc:
            errors.append({"row": index, "error": str(exc)})
        except (TypeError, ValueError, ValidationError) as exc:
            errors.append({"row": index, "error": str(exc) or "Invalid data in this row"})
        except Exception as exc:
            errors.append({"row": index, "error": "Unable to import this row"})
    return ok({"created_count": len(created), "created": created, "errors": errors})


@api_view(["GET", "POST", "PUT", "DELETE"])
def question_bank(request):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER, ROLE_STUDENT])
    if request.method == "GET":
        query = QuestionBankQuestion.objects
        if user.role == ROLE_STUDENT:
            student = get_student_for_user(user)
            query = query(class_level=student.class_level) if student else query(class_level="__none__")
        elif user.role == ROLE_TEACHER:
            query = query(class_level__in=teacher_classes(user))
            subjects = teacher_subjects(user)
            if subjects:
                query = query(subject__in=subjects)
        for key in ["class_level", "subject", "chapter", "difficulty", "question_type"]:
            value = request.GET.get(key)
            if value:
                query = query(**{key: value})
        search = str(request.GET.get("search", "")).strip()
        if search:
            query = query(text__icontains=search)
        return ok({"results": [question_bank_json(row, include_correct=user.role != ROLE_STUDENT) for row in query.order_by("-created_at")[:250]]})
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    data = request.data
    if request.method in ["PUT", "DELETE"]:
        row = QuestionBankQuestion.objects(id=data.get("id") or request.GET.get("id")).first()
        if not row:
            return bad("Question not found", status.HTTP_404_NOT_FOUND)
        enforce_teacher_question_scope(user, row.class_level, row.subject)
        if request.method == "DELETE":
            row.delete()
            return ok({"message": "Question deleted"})
        try:
            payload = question_bank_from_data(data, user, row)
        except (TypeError, ValueError):
            return bad("Unable to save this question.")
        row.update(**{f"set__{key}": value for key, value in payload.items()}, set__updated_at=datetime.utcnow())
        return ok({"question": question_bank_json(QuestionBankQuestion.objects(id=row.id).first(), include_correct=True)})
    try:
        payload = question_bank_from_data(data, user)
    except (TypeError, ValueError):
        return bad("Unable to create this question.")
    row = QuestionBankQuestion(**payload, created_by=user).save()
    return ok({"question": question_bank_json(row, include_correct=True)}, status.HTTP_201_CREATED)


@api_view(["GET"])
def daily_practice(request):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    session = get_or_create_daily_session(student)
    return ok({"session": practice_session_json(session, include_correct=session.status == "submitted")})


@api_view(["POST"])
def practice_answer(request, session_id):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    session = PracticeSession.objects(id=session_id, student=student).first()
    if not session:
        return bad("Unable to load this practice.", status.HTTP_404_NOT_FOUND)
    if session.status != "in_progress":
        return bad("This practice has already been submitted.", status.HTTP_403_FORBIDDEN)
    question = QuestionBankQuestion.objects(id=request.data.get("question_id")).first()
    if not question or oid(question) not in {oid(item) for item in session.questions}:
        return bad("Unable to save your answer.")
    answers = [answer for answer in session.answers if oid(answer.question) != oid(question)]
    answers.append(PracticeAnswer(question=question, answer=request.data.get("answer", ""), answered_at=datetime.utcnow()))
    session.update(answers=answers, updated_at=datetime.utcnow())
    return ok({"session": practice_session_json(PracticeSession.objects(id=session.id).first())})


@api_view(["POST"])
def practice_submit(request, session_id):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    session = PracticeSession.objects(id=session_id, student=student).first()
    if not session:
        return bad("Unable to load this practice.", status.HTTP_404_NOT_FOUND)
    session = submit_practice_session(session)
    return ok({"message": "Your practice has been submitted.", "session": practice_session_json(session, include_correct=True)})


@api_view(["GET"])
def practice_mistakes(request):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    rows = StudentMistake.objects(student=student).order_by("-last_wrong_at")
    return ok({"results": [mistake_json(row) for row in rows]})


@api_view(["POST"])
def practice_again(request, mistake_id):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    mistake = StudentMistake.objects(id=mistake_id, student=student).first()
    if not mistake or not mistake.question:
        return bad("Unable to load this mistake.", status.HTTP_404_NOT_FOUND)
    answer = str(request.data.get("answer", "")).strip()
    if not answer:
        return bad("Enter an answer to retry this question.")
    now = datetime.utcnow()
    correct = evaluate_practice_answer(mistake.question, answer)
    update_practice_tracking(student, mistake.question, correct, "practice", answer)
    refreshed = StudentMistake.objects(id=mistake.id, student=student).first()
    retry_attempts = getattr(refreshed, "retry_attempts", 0) + 1
    refreshed.update(
        retry_attempts=retry_attempts,
        last_retry_at=now,
        last_retry_answer=answer,
        last_retry_correct=correct,
        resolved=correct or refreshed.resolved,
        correct_streak=refreshed.correct_streak if correct else 0,
        corrected_at=now if correct else refreshed.corrected_at,
        updated_at=now,
    )
    return ok({"mistake": mistake_json(StudentMistake.objects(id=mistake.id, student=student).first()), "correct": correct})


@api_view(["GET"])
def weak_topics(request):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER, ROLE_STUDENT])
    query = TopicPerformance.objects
    if user.role == ROLE_STUDENT:
        query = query(student=get_student_for_user(user))
    elif user.role == ROLE_TEACHER:
        query = query(class_level__in=teacher_classes(user))
        subjects = teacher_subjects(user)
        if subjects:
            query = query(subject__in=subjects)
    for key in ["class_level", "subject", "chapter", "status"]:
        value = request.GET.get(key)
        if value:
            query = query(**{key: value})
    return ok({"results": [topic_performance_json(row) for row in query.order_by("accuracy", "-attempts")[:250]]})


@api_view(["GET"])
def study_plan(request):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    return ok({"plan": study_plan_json(get_or_create_study_plan(student))})


@api_view(["POST"])
def study_plan_task_complete(request, task_id):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    plan = StudyPlan.objects(student=student, plan_date=today_key()).first()
    if not plan:
        return bad("Study plan not found", status.HTTP_404_NOT_FOUND)
    requested_status = request.data.get("status")
    if requested_status not in ["Pending", "Completed", "Skipped", None]:
        return bad("Invalid task status.")
    completed = bool_value(request.data.get("completed", requested_status == "Completed"))
    tasks = []
    now = datetime.utcnow()
    found = False
    for task in plan.tasks:
        if task.task_id == task_id:
            found = True
            task.status = requested_status or ("Completed" if completed else "Pending")
            task.completed = completed
            task.completed_at = now if completed else None
        tasks.append(task)
    if not found:
        return bad("Study task not found", status.HTTP_404_NOT_FOUND)
    plan.update(tasks=tasks, updated_at=now)
    return ok({"plan": study_plan_json(StudyPlan.objects(id=plan.id).first())})


@api_view(["GET"])
def practice_analytics(request):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    query = TopicPerformance.objects
    if user.role == ROLE_TEACHER:
        query = query(class_level__in=teacher_classes(user))
        subjects = teacher_subjects(user)
        if subjects:
            query = query(subject__in=subjects)
    rows = list(query)
    if request.GET.get("class_level"):
        rows = [row for row in rows if row.class_level == request.GET.get("class_level")]
    class_map = {}
    subject_map = {}
    topic_rows = []
    for row in rows:
        for bucket, key in [(class_map, row.class_level), (subject_map, row.subject)]:
            current = bucket.get(key) or {"attempts": 0, "correct": 0}
            current["attempts"] += row.attempts
            current["correct"] += row.correct
            bucket[key] = current
        topic_rows.append(topic_performance_json(row))
    def mapped(data):
        return [
            {"label": key, "attempts": value["attempts"], "accuracy": round((value["correct"] / value["attempts"]) * 100, 2) if value["attempts"] else 0}
            for key, value in data.items()
        ]
    sessions = PracticeSession.objects
    if user.role == ROLE_TEACHER:
        sessions = sessions(class_level__in=teacher_classes(user))
    return ok(
        {
            "analytics": {
                "class_performance": sorted(mapped(class_map), key=lambda item: item["label"]),
                "subject_performance": sorted(mapped(subject_map), key=lambda item: item["accuracy"]),
                "difficult_topics": sorted(topic_rows, key=lambda item: (item["accuracy"], -item["attempts"]))[:10],
                "practice_participation": sessions.count(),
            }
        }
    )


def bool_value(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "published"}
    return bool(value)


def exam_window_state(exam, attempt=None):
    now = schedule_now()
    if attempt and attempt.status in ["submitted", "evaluated", "expired"]:
        return "completed"
    if not exam.is_published:
        return "draft"
    if now < schedule_time(exam.start_time):
        return "upcoming"
    if now > schedule_time(exam.end_time):
        return "ended"
    return "active"


def enforce_exam_manager(user, exam):
    require_roles_for_user(user, [ROLE_ADMIN, ROLE_TEACHER])
    enforce_teacher_class(user, exam.class_level)


def require_roles_for_user(user, roles):
    if user.role not in roles:
        raise PermissionDenied("You do not have permission to perform this action")


def parse_exam_datetime(value, field_label):
    try:
        return parse_date(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_label} must be a valid date and time.") from exc


def exam_fields(data, row=None):
    start_time = parse_exam_datetime(data.get("start_time", row.start_time if row else None), "Start time")
    end_time = parse_exam_datetime(data.get("end_time", row.end_time if row else None), "End time")
    if not start_time or not end_time:
        raise ValueError("Start time and end time are required.")
    if end_time <= start_time:
        raise ValueError("End time must be after start time.")
    try:
        duration = int(data.get("duration_minutes", row.duration_minutes if row else 60))
    except (TypeError, ValueError) as exc:
        raise ValueError("Duration must be a valid number of minutes.") from exc
    if duration < 1:
        raise ValueError("Duration must be at least 1 minute.")
    try:
        total_marks = float(data.get("total_marks", row.total_marks if row else 0))
        passing_marks = float(data.get("passing_marks", row.passing_marks if row else 0) or 0)
    except (TypeError, ValueError) as exc:
        raise ValueError("Marks must be valid numbers.") from exc
    return {
        "name": data.get("name", row.name if row else ""),
        "class_level": str(data.get("class_level", row.class_level if row else "")),
        "subject": data.get("subject", row.subject if row else ""),
        "instructions": data.get("instructions", row.instructions if row else ""),
        "exam_date": parse_exam_datetime(data.get("exam_date", row.exam_date if row else start_time), "Exam date"),
        "start_time": start_time,
        "end_time": end_time,
        "duration_minutes": duration,
        "total_marks": total_marks,
        "passing_marks": passing_marks,
        "is_published": bool_value(data.get("is_published", row.is_published if row else False)),
    }


def question_from_data(data, existing=None, order=0):
    qtype = data.get("question_type", existing.question_type if existing else "mcq")
    if qtype not in ["mcq", "true_false", "short", "long"]:
        raise ValueError("Invalid question type")
    options = data.get("options", existing.options if existing else [])
    if qtype == "mcq":
        options = [str(item).strip() for item in options if str(item).strip()][:4]
        if len(options) < 2:
            raise ValueError("MCQ questions need at least two options")
    elif qtype == "true_false":
        options = ["True", "False"]
    else:
        options = []
    text = data.get("text", existing.text if existing else "").strip()
    if not text:
        raise ValueError("Question text is required")
    try:
        marks = float(data.get("marks", existing.marks if existing else 1))
    except (TypeError, ValueError) as exc:
        raise ValueError("Marks must be greater than 0") from exc
    if marks <= 0:
        raise ValueError("Marks must be greater than 0")
    correct_answer = data.get("correct_answer", existing.correct_answer if existing else "").strip()
    expected_answer = data.get("expected_answer", existing.expected_answer if existing else "").strip()
    if qtype in ["mcq", "true_false"] and not correct_answer:
        raise ValueError("Correct answer missing")
    if qtype in ["short", "long"] and not expected_answer:
        raise ValueError("Expected answer missing")
    return ExamQuestion(
        question_id=existing.question_id if existing else uuid4().hex,
        question_bank_id=data.get("question_bank_id", existing.question_bank_id if existing else "").strip(),
        text=text,
        question_type=qtype,
        marks=marks,
        options=options,
        correct_answer=correct_answer,
        expected_answer=expected_answer,
        explanation=data.get("explanation", existing.explanation if existing else "").strip(),
        chapter=data.get("chapter", existing.chapter if existing else "").strip(),
        difficulty=data.get("difficulty", existing.difficulty if existing else "Medium"),
        order=int(data.get("order", existing.order if existing else order)),
    )


def question_json(question, include_correct=False):
    data = {
        "id": question.question_id,
        "question_id": question.question_id,
        "text": question.text,
        "question_type": question.question_type,
        "marks": question.marks,
        "options": question.options,
        "question_bank_id": question.question_bank_id,
        "chapter": question.chapter,
        "difficulty": question.difficulty,
        "explanation": question.explanation if include_correct else "",
        "expected_answer": question.expected_answer if include_correct else "",
        "order": question.order,
    }
    if include_correct:
        data["correct_answer"] = question.correct_answer
    return data


def answer_json(answer, include_marks=True):
    return {
        "question_id": answer.question_id,
        "answer": answer.answer,
        "marks_awarded": answer.marks_awarded if include_marks else 0,
        "feedback": answer.feedback if include_marks else "",
        "auto_graded": answer.auto_graded,
        "evaluated": answer.evaluated if include_marks else False,
    }


def attempt_json(attempt, include_answers=False, include_student=False, include_scores=True):
    data = {
        "id": oid(attempt),
        "exam_id": oid(attempt.exam),
        "started_at": dt(attempt.started_at),
        "deadline": dt(attempt.deadline),
        "submitted_at": dt(attempt.submitted_at),
        "status": attempt.status,
        "objective_score": attempt.objective_score if include_scores else 0,
        "descriptive_score": attempt.descriptive_score if include_scores else 0,
        "score": attempt.score if include_scores else 0,
        "feedback": attempt.feedback if include_scores else "",
        "violation_count": attempt.violation_count,
        "max_violations": settings.EXAM_MAX_SCREEN_VIOLATIONS,
        "remaining_violations": max(0, settings.EXAM_MAX_SCREEN_VIOLATIONS - attempt.violation_count),
        "last_violation_at": dt(attempt.last_violation_at),
        "auto_submitted": attempt.auto_submitted,
        "auto_submit_reason": attempt.auto_submit_reason,
    }
    if include_answers:
        data["answers"] = [answer_json(answer, include_marks=include_scores) for answer in attempt.answers]
    if include_student:
        data["student"] = student_json(attempt.student) if attempt.student else None
    return data


def exam_json(exam, user, include_questions=False, include_attempts=False):
    attempt = None
    if user.role == ROLE_STUDENT:
        student = get_student_for_user(user)
        attempt = ExamAttempt.objects(exam=exam, student=student).first() if student else None
    data = {
        "id": oid(exam),
        "name": exam.name,
        "class_level": exam.class_level,
        "subject": exam.subject,
        "instructions": exam.instructions,
        "exam_date": dt(exam.exam_date),
        "start_time": dt(exam.start_time),
        "end_time": dt(exam.end_time),
        "duration_minutes": exam.duration_minutes,
        "total_marks": exam.total_marks,
        "passing_marks": exam.passing_marks,
        "is_published": exam.is_published,
        "result_published": exam.result_published,
        "question_count": len(exam.questions),
        "status": exam_window_state(exam, attempt),
        "created_by": user_json(exam.created_by),
        "created_at": dt(exam.created_at),
        "updated_at": dt(exam.updated_at),
        "attempt": attempt_json(attempt, include_answers=True, include_scores=exam.result_published) if attempt else None,
    }
    if include_questions:
        data["questions"] = [question_json(question, include_correct=user.role != ROLE_STUDENT) for question in sorted(exam.questions, key=lambda item: item.order)]
    if include_attempts and user.role != ROLE_STUDENT:
        data["attempts"] = [attempt_json(row, include_answers=True, include_student=True) for row in ExamAttempt.objects(exam=exam).order_by("-updated_at")]
    return data


def visible_exams_for_user(user):
    if user.role == ROLE_STUDENT:
        student = get_student_for_user(user)
        return Exam.objects(class_level=student.class_level, is_published=True) if student else Exam.objects(class_level="__none__")
    if user.role == ROLE_TEACHER:
        return Exam.objects(class_level__in=teacher_classes(user))
    return Exam.objects


def auto_grade_attempt(attempt):
    exam = attempt.exam
    objective = 0
    by_id = {answer.question_id: answer for answer in attempt.answers}
    updated_answers = []
    for question in exam.questions:
        answer = by_id.get(question.question_id) or ExamAnswer(question_id=question.question_id)
        if question.question_type in ["mcq", "true_false"]:
            answer.auto_graded = True
            answer.evaluated = True
            answer.marks_awarded = question.marks if normalize_answer(answer.answer) == normalize_answer(objective_correct_answer(question)) else 0
            objective += answer.marks_awarded
        updated_answers.append(answer)
    attempt.update(
        answers=updated_answers,
        objective_score=objective,
        score=objective + attempt.descriptive_score,
        updated_at=datetime.utcnow(),
    )
    return ExamAttempt.objects(id=attempt.id).first()


def finalize_attempt(attempt, status_value="submitted", auto_submitted=False, auto_submit_reason=""):
    if attempt.status in ["submitted", "evaluated", "expired"]:
        return attempt
    now = store_schedule_time(schedule_now())
    update_fields = {
        "status": status_value,
        "submitted_at": now,
        "updated_at": datetime.utcnow(),
    }
    if auto_submitted:
        update_fields["auto_submitted"] = True
        update_fields["auto_submit_reason"] = auto_submit_reason
    attempt.update(**update_fields)
    return auto_grade_attempt(ExamAttempt.objects(id=attempt.id).first())


def anti_cheat_message(count):
    return "Your exam was automatically submitted because you left the exam screen."


def auto_submit_reason_from_request(reason):
    if reason == "fullscreen_exit":
        return "fullscreen_exit"
    return "tab_window_exit"


@api_view(["GET", "POST", "PUT", "DELETE"])
def exams(request):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER, ROLE_STUDENT])
    if request.method == "GET":
        query = visible_exams_for_user(user)
        for key in ["class_level", "subject"]:
            value = request.GET.get(key)
            if value:
                query = query(**{key: value})
        return ok({"server_time": dt(store_schedule_time(schedule_now())), "results": [exam_json(row, user, include_questions=user.role != ROLE_STUDENT, include_attempts=user.role != ROLE_STUDENT) for row in query.order_by("-start_time")]})
    require_roles_for_user(user, [ROLE_ADMIN, ROLE_TEACHER])
    data = request.data
    if request.method in ["PUT", "DELETE"]:
        row = Exam.objects(id=data.get("id") or request.GET.get("id")).first()
        if not row:
            return bad("Exam not found", status.HTTP_404_NOT_FOUND)
        enforce_exam_manager(user, row)
        if request.method == "DELETE":
            require_roles_for_user(user, [ROLE_ADMIN])
            row.delete()
            return ok({"message": "Exam deleted"})
        try:
            payload = exam_fields(data, row)
            enforce_teacher_class(user, payload["class_level"])
        except (TypeError, ValueError) as exc:
            return bad(str(exc) or "Unable to save this exam.")
        was_published = row.is_published
        row.update(**{f"set__{key}": value for key, value in payload.items()}, set__updated_at=datetime.utcnow())
        updated = Exam.objects(id=row.id).first()
        if not was_published and updated.is_published:
            notify_students_for_class(updated.class_level, "exam", "Exam Published", f"{updated.name} is now available.", "/operations", "red", "assignment", updated.id)
        return ok({"exam": exam_json(updated, user, include_questions=True, include_attempts=True)})
    try:
        payload = exam_fields(data)
        enforce_teacher_class(user, payload["class_level"])
    except (TypeError, ValueError) as exc:
        return bad(str(exc) or "Unable to create this exam.")
    row = Exam(**payload, created_by=user).save()
    if row.is_published:
        notify_students_for_class(row.class_level, "exam", "Exam Published", f"{row.name} is now available.", "/operations", "red", "assignment", row.id)
    return ok({"exam": exam_json(row, user, include_questions=True, include_attempts=True)}, status.HTTP_201_CREATED)


@api_view(["POST", "PUT", "DELETE"])
def exam_questions(request, exam_id):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    exam = Exam.objects(id=exam_id).first()
    if not exam:
        return bad("Exam not found", status.HTTP_404_NOT_FOUND)
    enforce_exam_manager(user, exam)
    data = request.data
    questions = list(exam.questions)
    if request.method == "DELETE":
        qid = data.get("question_id") or request.GET.get("question_id")
        questions = [question for question in questions if question.question_id != qid]
    elif request.method == "PUT":
        qid = data.get("question_id")
        existing = next((question for question in questions if question.question_id == qid), None)
        if not existing:
            return bad("Question not found", status.HTTP_404_NOT_FOUND)
        try:
            updated = question_from_data(data, existing)
        except (TypeError, ValueError):
            return bad("Unable to save this question.")
        questions = [updated if question.question_id == qid else question for question in questions]
    else:
        bank_ids = data.get("question_bank_ids") or []
        if bank_ids:
            existing_bank_ids = {question.question_bank_id for question in questions if question.question_bank_id}
            for bank_question in QuestionBankQuestion.objects(id__in=bank_ids):
                if bank_question.class_level != exam.class_level or bank_question.subject != exam.subject or oid(bank_question) in existing_bank_ids:
                    continue
                enforce_teacher_question_scope(user, bank_question.class_level, bank_question.subject)
                questions.append(exam_question_from_bank(bank_question, len(questions) + 1))
                existing_bank_ids.add(oid(bank_question))
        else:
            try:
                questions.append(question_from_data(data, order=len(questions) + 1))
            except (TypeError, ValueError):
                return bad("Unable to add this question.")
    questions = sorted(questions, key=lambda question: question.order)
    exam.update(questions=questions, total_marks=sum(question.marks for question in questions), updated_at=datetime.utcnow())
    return ok({"exam": exam_json(Exam.objects(id=exam.id).first(), user, include_questions=True, include_attempts=True)})


EXAM_QUESTION_BULK_HEADERS = [
    "Question",
    "Type",
    "Option A",
    "Option B",
    "Option C",
    "Option D",
    "Correct Answer",
    "Expected Answer",
    "Marks",
    "Chapter",
    "Difficulty",
    "Explanation",
]


def exam_question_template(request):
    require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(EXAM_QUESTION_BULK_HEADERS)
    writer.writerow(["What is 2 + 2?", "mcq", "2", "3", "4", "5", "4", "", "1", "Numbers", "Easy", "Basic addition"])
    writer.writerow(["The Sun rises in the east.", "true_false", "", "", "", "", "True", "", "1", "General", "Easy", ""])
    writer.writerow(["Explain photosynthesis.", "short", "", "", "", "", "", "Plants make food using sunlight.", "3", "Biology", "Medium", ""])
    response = HttpResponse(output.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="exam-question-template.csv"'
    return response


def uploaded_question_rows(upload):
    name = (upload.name or "").lower()
    if name.endswith(".xlsx"):
        workbook = load_workbook(upload, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        return [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:]
            if any(value not in [None, ""] for value in row)
        ]
    text = upload.read().decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def row_value(row, *keys):
    lowered = {str(key).strip().lower(): value for key, value in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value not in [None, ""]:
            return str(value).strip()
    return ""


def exam_question_data_from_row(row):
    qtype_raw = row_value(row, "Type", "question_type").lower() or "mcq"
    qtype = {"true/false": "true_false", "true false": "true_false", "tf": "true_false"}.get(qtype_raw, qtype_raw)
    expected = row_value(row, "Expected Answer", "expected_answer")
    correct = row_value(row, "Correct Answer", "correct_answer")
    return {
        "text": row_value(row, "Question", "question", "text"),
        "question_type": qtype,
        "options": [row_value(row, "Option A", "option_a"), row_value(row, "Option B", "option_b"), row_value(row, "Option C", "option_c"), row_value(row, "Option D", "option_d")],
        "correct_answer": correct,
        "expected_answer": expected or (correct if qtype in ["short", "long"] else ""),
        "marks": row_value(row, "Marks", "marks") or 1,
        "chapter": row_value(row, "Chapter", "chapter", "topic"),
        "difficulty": row_value(row, "Difficulty", "difficulty") or "Medium",
        "explanation": row_value(row, "Explanation", "explanation"),
    }


@api_view(["POST"])
def exam_questions_bulk_import(request, exam_id):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    exam = Exam.objects(id=exam_id).first()
    if not exam:
        return bad("Exam not found", status.HTTP_404_NOT_FOUND)
    enforce_exam_manager(user, exam)
    upload = request.FILES.get("file")
    if not upload:
        return bad("Upload a CSV or XLSX file with exam questions.")
    try:
        rows = uploaded_question_rows(upload)
    except UnicodeDecodeError:
        return bad("Unable to read this file. Please upload a CSV or XLSX file.")
    except Exception:
        return bad("Unable to read this file. Please upload a CSV or XLSX file.")
    existing_texts = {question.text.strip().lower() for question in exam.questions}
    seen_texts = set()
    parsed_questions = []
    errors = []
    for index, row in enumerate(rows, start=2):
        try:
            body = exam_question_data_from_row(row)
            text_key = body["text"].strip().lower()
            if not text_key:
                raise ValueError("Question text is required")
            if text_key in existing_texts or text_key in seen_texts:
                raise ValueError("Duplicate question")
            parsed_questions.append(question_from_data(body, order=len(exam.questions) + len(parsed_questions) + 1))
            seen_texts.add(text_key)
        except (TypeError, ValueError, ValidationError) as exc:
            errors.append({"row": index, "error": str(exc) or "Invalid data in this row"})
    if errors:
        return ok({"total_rows": len(rows), "valid_count": len(parsed_questions), "invalid_count": len(errors), "errors": errors}, status.HTTP_400_BAD_REQUEST)
    questions = sorted(list(exam.questions) + parsed_questions, key=lambda question: question.order)
    exam.update(questions=questions, total_marks=sum(question.marks for question in questions), updated_at=datetime.utcnow())
    return ok(
        {
            "total_rows": len(rows),
            "valid_count": len(parsed_questions),
            "invalid_count": 0,
            "errors": [],
            "exam": exam_json(Exam.objects(id=exam.id).first(), user, include_questions=True, include_attempts=True),
        },
        status.HTTP_201_CREATED,
    )


@api_view(["POST"])
def exam_duplicate(request, exam_id):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    exam = Exam.objects(id=exam_id).first()
    if not exam:
        return bad("Exam not found", status.HTTP_404_NOT_FOUND)
    enforce_exam_manager(user, exam)
    questions = []
    for index, question in enumerate(sorted(exam.questions, key=lambda item: item.order), start=1):
        questions.append(
            ExamQuestion(
                question_id=uuid4().hex,
                question_bank_id=question.question_bank_id,
                text=question.text,
                question_type=question.question_type,
                marks=question.marks,
                options=list(question.options),
                correct_answer=question.correct_answer,
                expected_answer=question.expected_answer,
                explanation=question.explanation,
                chapter=question.chapter,
                difficulty=question.difficulty,
                order=index,
            )
        )
    row = Exam(
        name=f"{exam.name} Copy",
        class_level=exam.class_level,
        subject=exam.subject,
        instructions=exam.instructions,
        exam_date=exam.exam_date,
        start_time=exam.start_time,
        end_time=exam.end_time,
        duration_minutes=exam.duration_minutes,
        total_marks=exam.total_marks,
        passing_marks=exam.passing_marks,
        is_published=False,
        result_published=False,
        questions=questions,
        created_by=user,
    ).save()
    return ok({"exam": exam_json(row, user, include_questions=True, include_attempts=True)}, status.HTTP_201_CREATED)


@api_view(["POST"])
def exam_start(request, exam_id):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    exam = Exam.objects(id=exam_id, is_published=True).first()
    if not exam or not student or exam.class_level != student.class_level:
        return bad("Unable to load this exam.", status.HTTP_404_NOT_FOUND)
    now = schedule_now()
    start_time = schedule_time(exam.start_time)
    end_time = schedule_time(exam.end_time)
    if now < start_time:
        return bad("This exam has not started yet.", status.HTTP_403_FORBIDDEN)
    if now > end_time:
        return bad("This exam has ended.", status.HTTP_403_FORBIDDEN)
    attempt = ExamAttempt.objects(exam=exam, student=student).first()
    if attempt:
        if attempt.status == "in_progress" and now > schedule_time(attempt.deadline):
            attempt = finalize_attempt(attempt, "expired")
        return ok({"exam": exam_json(exam, user, include_questions=True), "attempt": attempt_json(attempt, include_answers=True, include_scores=exam.result_published), "server_time": dt(store_schedule_time(now))})
    deadline = min(now + timedelta(minutes=exam.duration_minutes), end_time)
    attempt = ExamAttempt(exam=exam, student=student, started_at=store_schedule_time(now), deadline=store_schedule_time(deadline)).save()
    return ok({"exam": exam_json(exam, user, include_questions=True), "attempt": attempt_json(attempt, include_answers=True, include_scores=exam.result_published), "server_time": dt(store_schedule_time(now))}, status.HTTP_201_CREATED)


@api_view(["POST"])
def exam_attempt_answer(request, attempt_id):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    attempt = ExamAttempt.objects(id=attempt_id, student=student).first()
    if not attempt:
        return bad("Unable to load this exam.", status.HTTP_404_NOT_FOUND)
    now = schedule_now()
    if attempt.status != "in_progress" or now > schedule_time(attempt.deadline) or now > schedule_time(attempt.exam.end_time):
        finalize_attempt(attempt, "expired" if attempt.status == "in_progress" else attempt.status)
        return bad("This exam has ended.", status.HTTP_403_FORBIDDEN)
    qid = request.data.get("question_id")
    answer_text = request.data.get("answer", "")
    question_ids = {question.question_id for question in attempt.exam.questions}
    if qid not in question_ids:
        return bad("Unable to save your answer.")
    answers = [answer for answer in attempt.answers if answer.question_id != qid]
    answers.append(ExamAnswer(question_id=qid, answer=answer_text))
    attempt.update(answers=answers, updated_at=datetime.utcnow())
    return ok({"attempt": attempt_json(ExamAttempt.objects(id=attempt.id).first(), include_answers=True, include_scores=attempt.exam.result_published)})


@api_view(["POST"])
def exam_attempt_submit(request, attempt_id):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    attempt = ExamAttempt.objects(id=attempt_id, student=student).first()
    if not attempt:
        return bad("Unable to load this exam.", status.HTTP_404_NOT_FOUND)
    now = schedule_now()
    status_value = "expired" if now > schedule_time(attempt.deadline) or now > schedule_time(attempt.exam.end_time) else "submitted"
    attempt = finalize_attempt(attempt, status_value)
    return ok({"message": "Your exam has been submitted successfully.", "attempt": attempt_json(attempt, include_answers=True, include_scores=attempt.exam.result_published)})


@api_view(["POST"])
def exam_attempt_violation(request, attempt_id):
    user = require_roles(request, [ROLE_STUDENT])
    student = get_student_for_user(user)
    attempt = ExamAttempt.objects(id=attempt_id, student=student).first()
    if not attempt:
        return bad("Unable to load this exam.", status.HTTP_404_NOT_FOUND)
    now = schedule_now()
    if attempt.status != "in_progress":
        return ok({"message": "This exam has already been submitted.", "attempt": attempt_json(attempt, include_answers=True, include_scores=attempt.exam.result_published), "auto_submitted": attempt.auto_submitted})
    if now > schedule_time(attempt.deadline) or now > schedule_time(attempt.exam.end_time):
        attempt = finalize_attempt(attempt, "expired")
        return ok({"message": "Time is over. Your exam has been submitted.", "attempt": attempt_json(attempt, include_answers=True, include_scores=attempt.exam.result_published), "auto_submitted": False})
    updated = ExamAttempt.objects(id=attempt.id, student=student, status="in_progress").modify(
        new=True,
        inc__violation_count=1,
        set__last_violation_at=store_schedule_time(now),
        set__updated_at=datetime.utcnow(),
    )
    if not updated:
        current = ExamAttempt.objects(id=attempt.id, student=student).first()
        return ok({"message": "This exam has already been submitted.", "attempt": attempt_json(current, include_answers=True, include_scores=current.exam.result_published), "auto_submitted": current.auto_submitted})
    updated = finalize_attempt(updated, "submitted", auto_submitted=True, auto_submit_reason=auto_submit_reason_from_request(request.data.get("reason", "")))
    return ok({"message": anti_cheat_message(updated.violation_count), "attempt": attempt_json(updated, include_answers=True, include_scores=updated.exam.result_published), "auto_submitted": True})


@api_view(["PUT"])
def exam_attempt_evaluate(request, attempt_id):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    attempt = ExamAttempt.objects(id=attempt_id).first()
    if not attempt:
        return bad("Submission not found", status.HTTP_404_NOT_FOUND)
    enforce_exam_manager(user, attempt.exam)
    evaluations = {item.get("question_id"): item for item in request.data.get("evaluations", [])}
    answers = []
    descriptive = 0
    objective = 0
    questions = {question.question_id: question for question in attempt.exam.questions}
    existing = {answer.question_id: answer for answer in attempt.answers}
    for qid, question in questions.items():
        answer = existing.get(qid) or ExamAnswer(question_id=qid)
        if question.question_type in ["mcq", "true_false"]:
            objective += answer.marks_awarded
        else:
            evaluation = evaluations.get(qid, {})
            marks_value = min(float(evaluation.get("marks_awarded", answer.marks_awarded) or 0), question.marks)
            answer.marks_awarded = max(0, marks_value)
            answer.feedback = evaluation.get("feedback", answer.feedback)
            answer.evaluated = True
            descriptive += answer.marks_awarded
        answers.append(answer)
    attempt.update(
        answers=answers,
        objective_score=objective,
        descriptive_score=descriptive,
        score=objective + descriptive,
        status="evaluated",
        feedback=request.data.get("feedback", attempt.feedback),
        evaluated_by=user,
        evaluated_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    return ok({"attempt": attempt_json(ExamAttempt.objects(id=attempt.id).first(), include_answers=True, include_student=True)})


@api_view(["POST"])
def exam_publish_results(request, exam_id):
    user = require_roles(request, [ROLE_ADMIN, ROLE_TEACHER])
    exam = Exam.objects(id=exam_id).first()
    if not exam:
        return bad("Exam not found", status.HTTP_404_NOT_FOUND)
    enforce_exam_manager(user, exam)
    exam.update(result_published=True, updated_at=datetime.utcnow())
    process_published_exam_attempts(exam)
    notify_students_for_class(exam.class_level, "exam_result", "Exam Result Published", f"{exam.name} result is now available.", "/operations", "green", "marks", exam.id)
    return ok({"exam": exam_json(Exam.objects(id=exam.id).first(), user, include_questions=True, include_attempts=True)})
